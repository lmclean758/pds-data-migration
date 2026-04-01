#!/usr/bin/env python3
"""
PDS File Scanner
================
Scans a folder of PDS Excel files and populates PDS_TRACKER_FILE.xlsx.

How it works
------------
- Row 6 of the tracker's "Packaging Proposal Template" sheet contains
  instructions for each column, colour-coded:
    Orange (FFFFC000) : Fixed literal value (use the text in the cell as-is)
    Yellow (FFFFFF00) : Leave blank
    Green  (FF92D050) : Read from the PDS cell whose address is in the cell
    Red    (FFFF0000) : Evaluate the arithmetic formula described in the cell

- The Part Number column (M) has instruction "K16 - new line for each PN".
  In the PDS there are 12 merged-cell regions (K16:AM17 area) that each
  hold one part number. One tracker row is written per non-empty part number,
  with all other data identical (same PDS → same packaging / dimensions).

- Rows that already exist in the tracker (matched on Part Number in column M)
  are skipped so you can re-run the script safely.
"""

import os
import re
import glob
import shutil
import warnings
from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
PDS_FOLDER = r'C:\Users\LukeMclean\Desktop\PDS_FILES_ALL'  # ADJUST THIS PATH
TRACKER_FILE = r'C:\Users\LukeMclean\Desktop\PDS_TRACKER_FILE.xlsx'
OUTPUT_FOLDER = r'C:\Users\LukeMclean\Desktop'

TRACKER_SHEET = 'Packaging Proposal Template'
INSTRUCTION_ROW = 6          # Row in tracker that holds the mapping instructions
HEADER_ROWS = 6          # Rows to skip when looking for existing data

# The 12 merged-cell "top-left" cells that may contain a part number
PART_NUMBER_CELLS = [
    'K16', 'M16', 'S16', 'Z16', 'AG16', 'AM16',
    'K17', 'M17', 'S17', 'Z17', 'AG17', 'AM17',
]

# Colour hex values (ARGB) that appear in instruction row 6
COLOR_YELLOW = 'FFFFFF00'   # Blank
COLOR_ORANGE = 'FFFFC000'   # Fixed value
COLOR_GREEN = 'FF92D050'   # Read from PDS cell
COLOR_RED = 'FFFF0000'   # Calculation

# ── Helpers ───────────────────────────────────────────────────────────────────


def _cell_bg(cell) -> str | None:
    """Return the ARGB background colour of a cell, or None."""
    try:
        rgb = cell.fill.fgColor.rgb
        return rgb if rgb not in (None, '00000000', 'FF000000') else None
    except Exception:
        return None


def _safe_float(value):
    """Convert *value* to float; return None on failure."""
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _pds_sheet_name(wb) -> str:
    """Return the name of the PDS data sheet (stripped to 'PDS', case-insensitive)."""
    for name in wb.sheetnames:
        if name.strip().upper() == 'PDS':
            return name
    return wb.sheetnames[0]


def load_pds_values(pds_path: str) -> dict:
    """
    Load all cell values from the PDS sheet into a plain dict
    { 'K16': value, 'M27': value, ... } using a single sequential pass
    in read_only mode.  This is fast and skips all style/border processing
    that can crash older openpyxl versions.
    """
    wb = openpyxl.load_workbook(pds_path, data_only=True, read_only=True)
    sheet_name = _pds_sheet_name(wb)
    ws = wb[sheet_name]
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                ref = f"{get_column_letter(cell.column)}{cell.row}"
                cells[ref] = cell.value
    wb.close()
    return cells


# ── Instruction parsing ───────────────────────────────────────────────────────

def read_instructions(tracker_ws) -> dict:
    """
    Parse row 6 of the tracker into a mapping:
        { 'B': {'type': 'blank'}, 'C': {'type': 'fixed', 'value': 'Approved'}, ... }
    """
    instructions = {}
    for cell in tracker_ws[INSTRUCTION_ROW]:
        col = get_column_letter(cell.column)
        bg = _cell_bg(cell)
        val = cell.value

        if bg == COLOR_YELLOW:
            instructions[col] = {'type': 'blank'}
        elif bg == COLOR_ORANGE:
            instructions[col] = {'type': 'fixed', 'value': val}
        elif bg == COLOR_GREEN:
            text = str(val).strip() if val else ''
            if 'new line for each PN' in text:
                instructions[col] = {'type': 'part_number'}
            else:
                instructions[col] = {'type': 'pds', 'cell': text}
        elif bg == COLOR_RED:
            instructions[col] = {'type': 'calc', 'formula': str(val) if val else ''}

    return instructions


# ── Calculation evaluator ─────────────────────────────────────────────────────

# Matches uppercase cell references like K19, AA26, AK40, etc.
_CELL_REF_RE = re.compile(r'\b([A-Z]{1,3}\d{1,3})\b')


def _evaluate_formula(formula_text: str, pds_cells: dict) -> float | None:
    """
    Evaluate a red-cell formula.  Examples of *formula_text*:
        'This equals cells W40 - K19'
        'This equals cells AK40 - (K19 * AA26)'
        'This equals cells AT35 * BC35'
    Returns a float or None if any referenced cell is empty.
    """
    # Extract the arithmetic portion after 'This equals cells'
    m = re.search(r'[Tt]his equals cells?\s+(.+)', formula_text)
    if not m:
        return None

    expr = m.group(1).strip()

    # Collect all unique cell references (longest first to avoid partial hits)
    refs = sorted(set(_CELL_REF_RE.findall(expr)), key=len, reverse=True)

    for ref in refs:
        v = _safe_float(pds_cells.get(ref))
        if v is None:
            return None
        # Replace whole-word occurrences only
        expr = re.sub(rf'\b{re.escape(ref)}\b', str(v), expr)

    # Allow only safe arithmetic characters before eval
    if re.fullmatch(r'[\d\s\+\-\*\/\(\)\.]+', expr):
        try:
            return eval(expr)   # noqa: S307 — expression is sanitised above
        except Exception:
            return None

    return None


# ── Part-number extraction ────────────────────────────────────────────────────

def get_part_numbers(pds_cells: dict) -> list[str]:
    """Return all non-empty part numbers from the 12 candidate cells."""
    pns = []
    for ref in PART_NUMBER_CELLS:
        val = pds_cells.get(ref)
        if val is not None and str(val).strip():
            pns.append(str(val).strip())
    return pns


# ── Row building ──────────────────────────────────────────────────────────────

def build_row_data(pds_cells: dict, instructions: dict, part_number: str) -> dict:
    """
    Build a complete row dict { col_letter: value } for one part number.
    All data (except the part number itself) come from *pds_cells*.
    """
    row = {}
    for col, inst in instructions.items():
        t = inst['type']
        if t == 'blank':
            row[col] = None
        elif t == 'fixed':
            row[col] = inst['value']
        elif t == 'pds':
            row[col] = pds_cells.get(inst['cell']) if inst['cell'] else None
        elif t == 'calc':
            row[col] = _evaluate_formula(inst['formula'], pds_cells)
        elif t == 'part_number':
            row[col] = part_number
    return row


# ── Tracker helpers ───────────────────────────────────────────────────────────

def find_part_number_column(instructions: dict) -> str | None:
    """Return the tracker column letter used for Part Number."""
    for col, inst in instructions.items():
        if inst['type'] == 'part_number':
            return col
    return None


def existing_part_numbers(tracker_ws, pn_col: str) -> set:
    """
    Return the set of part numbers already in the tracker
    (all rows below the instruction row).
    """
    col_idx = column_index_from_string(pn_col)
    pns = set()
    for row in tracker_ws.iter_rows(
        min_row=INSTRUCTION_ROW + 1, min_col=col_idx, max_col=col_idx
    ):
        val = row[0].value
        if val is not None:
            pns.add(str(val).strip())
    return pns


def next_empty_row(tracker_ws) -> int:
    """
    Return the first row (>= INSTRUCTION_ROW+1) where columns C onward are
    all empty — i.e. the next row available for new data.
    Column B is excluded because it holds the legend in rows 12-15.
    """
    last_data = INSTRUCTION_ROW
    for row in tracker_ws.iter_rows(min_row=INSTRUCTION_ROW + 1):
        for cell in row:
            if cell.column > 2 and cell.value is not None:
                if cell.row > last_data:
                    last_data = cell.row
    return last_data + 1


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # Suppress noisy openpyxl warnings about unsupported extensions / drawings
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

    # Build a dated output path and copy the tracker as the template
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(OUTPUT_FOLDER, f'PDS_OUTPUT_{timestamp}.xlsx')
    shutil.copy2(TRACKER_FILE, output_file)
    print(f"Output file : {output_file}")

    print(f"Loading template: {TRACKER_FILE}")
    tracker_wb = openpyxl.load_workbook(output_file)
    tracker_ws = tracker_wb[TRACKER_SHEET]

    instructions = read_instructions(tracker_ws)
    pn_col = find_part_number_column(instructions)

    if pn_col is None:
        print("ERROR: Could not find the Part Number column in the tracker row 6.")
        return

    print(f"Part Number column in tracker: {pn_col}")

    already_present = existing_part_numbers(tracker_ws, pn_col)
    write_row = next_empty_row(tracker_ws)

    # Unmerge any merged ranges that overlap the data area so we can write
    # freely.  Use max_row (not min_row) so ranges that START in the header
    # but EXTEND into data rows are also caught.
    data_merges = [
        str(mc) for mc in list(tracker_ws.merged_cells.ranges)
        if mc.max_row > INSTRUCTION_ROW
    ]
    for mc_ref in data_merges:
        tracker_ws.unmerge_cells(mc_ref)

    print(f"Writing new rows from row {write_row}\n")

    pds_files = sorted(glob.glob(os.path.join(PDS_FOLDER, '*.xlsx')))
    rows_added = 0
    skipped = 0
    errors = []

    for pds_path in pds_files:
        filename = os.path.basename(pds_path)
        # Skip Office temp lock files
        if filename.startswith('~$'):
            continue
        try:
            # read_only=True loads cells in a single pass — fast and skips all
            # style/border processing that causes an openpyxl bug on some files.
            pds_cells = load_pds_values(pds_path)

            part_numbers = get_part_numbers(pds_cells)

            if not part_numbers:
                print(f"  [SKIP]  {filename}  \u2014 no part numbers found")
                skipped += 1
                continue

            for pn in part_numbers:
                if pn in already_present:
                    print(f"  [DUP]   {pn}  \u2014 already in tracker, skipping")
                    skipped += 1
                    continue

                row_data = build_row_data(pds_cells, instructions, pn)

                for col_letter, value in row_data.items():
                    col_idx = column_index_from_string(col_letter)
                    tracker_ws.cell(row=write_row, column=col_idx, value=value)

                print(f"  [OK]    Row {write_row:>4d}  {pn:<30s}  ({filename})")
                already_present.add(pn)
                write_row += 1
                rows_added += 1

        except Exception as exc:
            msg = f"{filename}: {exc}"
            errors.append(msg)
            print(f"  [ERR]   {msg}")

    print("\nSaving output ...")
    tracker_wb.save(output_file)

    print(f"\n{'='*60}")
    print(f"  PDS files found : {len(pds_files)}")
    print(f"  Rows added      : {rows_added}")
    print(f"  Skipped / dup   : {skipped}")
    if errors:
        print(f"  Errors          : {len(errors)}")
        for e in errors:
            print(f"    • {e}")
    print(f"{'='*60}")
    print(f"Output saved to : {output_file}")


if __name__ == '__main__':
    main()
